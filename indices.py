#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Jan 25 15:41:45 2019

@author: Lau
"""
"""
import pandas as pd

df = pd.read_excel("datos.xlsx", header = None).pivot(columns = 0,  values = 1)
Lista = {col: df[col].dropna().tolist() for col in df.columns}
"""

import openpyxl

#Vamos a leer el documento excel para el índice AEX25
doc_aex = openpyxl.load_workbook('/DATOS_HISTORICOS_INDICES/AEX25/AEX.xls')

doc_aex.get_sheet_names()   #Con esto sacamos las hojas que tiene cada documento excel
hoja_aex = doc_aex.get_sheet_by_name('Hoja1')   #Con esto abrimos la Hoja1

for columna in hoja_aex.column=1:   #Sacamos los valores de la columna 1 y los metemos en una lista que serán las fechas del índice
    aex_fechas = columna

for columna in hoja_aex.column=2:   #Sacamos los valores de la columna 2 y los metemos en una lista que serán los precios de cierre del índice
    aex_cierre = columna

#Vamos a leer el documento excel para el índice CAC40
doc_cac = openpyxl.load_workbook('/DATOS_HISTORICOS_INDICES/CAC40/CAC40.xls')

doc_cac.get_sheet_names()   
hoja_cac = doc_cac.get_sheet_by_name('Hoja1')  

for columna in hoja_cac.column=1:
    cac_fechas = columna

for columna in hoja_cac.column=2:
    cac_cierre = columna

#Vamos a leer el documento excel para el índice DAX
doc_dax = openpyxl.load_workbook('/DATOS_HISTORICOS_INDICES/DAX/DAX.xls')

doc_dax.get_sheet_names()   
hoja_dax = doc_dax.get_sheet_by_name('Hoja1')   

for columna in hoja_dax.column=1:
    dax_fechas = columna

for columna in hoja_dax.column=2:
    dax_cierre = columna
    
#Vamos a leer el documento excel para el índice FTSE100
doc_ftse = openpyxl.load_workbook('/DATOS_HISTORICOS_INDICES/FTSE100/FTSE100.xls')

doc_ftse.get_sheet_names()
hoja_ftse = doc_ftse.get_sheet_by_name('Hoja1')  

for columna in hoja_ftse.column=1:
    ftse_fechas = columna

for columna in hoja_ftse.column=2:
    ftse_cierre = columna
    
#Vamos a leer el documento excel para el índice HSI
doc_hsi = openpyxl.load_workbook('/DATOS_HISTORICOS_INDICES/HSI/HSI.xls')

doc_hsi.get_sheet_names()   
hoja_hsi = dodoc_hsic.get_sheet_by_name('Hoja1')  

for columna in hoja_hsi.column=1:
    hsi_fechas = columna

for columna in hoja_hsi.column=2:
    hsi_cierre = columna 

#Vamos a leer el documento excel para el índice IBEX35
doc_ibex = openpyxl.load_workbook('/DATOS_HISTORICOS_INDICES/IBEX35/IBEX35.xls')

doc_ibex.get_sheet_names()
hoja_ibex = doc_ibex.get_sheet_by_name('Hoja1')

for columna in hoja_ibex.column=1:
    ibex_fechas = columna

for columna in hoja_ibex.column=2:
    ibex_cierre = columna

#Vamos a leer el documento excel para el índice NIFTY50
doc_nifty = openpyxl.load_workbook('/DATOS_HISTORICOS_INDICES/NIFTY50/NIFTY50.xls')

doc_nifty.get_sheet_names()
hoja_nifty = doc_nifty.get_sheet_by_name('Hoja1')

for columna in hoja_nifty.column=1:
    nifty_fechas = columna

for columna in hoja_nifty.column=2:
    nifty_cierre = columna

#Vamos a leer el documento excel para el índice NIKKEI225
doc_nikkei = openpyxl.load_workbook('/DATOS_HISTORICOS_INDICES/NIKKEI225/NIKKEI225.xls')

doc_nikkei.get_sheet_names()
hoja_nikkei = doc_nikkei.get_sheet_by_name('Hoja1')

for columna in hoja_nikkei.column=1:
    nikkei_fechas = columna

for columna in hoja_nikkei.column=2:
    nikkei_cierre = columna

#Vamos a leer el documento excel para el índice S&P500
doc_syp = openpyxl.load_workbook('/DATOS_HISTORICOS_INDICES/S&P500/S&P500.xls')

doc_syp.get_sheet_names()
hoja_syp = doc_syp.get_sheet_by_name('Hoja1')

for columna in hoja_syp.column=1:
    syp_fechas = columna

for columna in hoja_syp.column=2:
    syp_cierre = columna
    

#Vamos a leer el documento excel para el índice SSEC
doc_ssec = openpyxl.load_workbook('/DATOS_HISTORICOS_INDICES/SSEC/SSEC.xls')

doc_ssec.get_sheet_names()
hoja_ssec = doc_ssec.get_sheet_by_name('Hoja1')

for columna in hoja_ssec.column=1:
    ssec_fechas = columna

for columna in hoja_ssec.column=2:
    ssec_cierre = columna